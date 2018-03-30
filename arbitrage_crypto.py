import os
import sys
import datetime
import requests
import json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook
from openpyxl import load_workbook
from tabulate import tabulate


def convert_usd_to_inr():
	url = "https://api.fixer.io/latest?base=USD"
	r = requests.get(url)
	data = json.loads(r.text)
	
	return float(data['rates']['INR'])


def convert_usdt_to_usd():
	url = "https://api.exmo.com/v1/ticker"
	r = requests.get(url)
	data = json.loads(r.text)
	
	return float(data['USDT_USD']['last_trade'])

def convert_eth_to_usdt():
	url = "https://poloniex.com/public?command=returnTicker"
	r = requests.get(url)
	data = json.loads(r.text)
	
	return float(data['USDT_ETH']['last'])

def convert_eth_to_inr():
	url = "https://koinex.in/api/ticker"
	r = requests.get(url)
	data = json.loads(r.text)
	
	return float(data['prices']['ETH'])



#Initialize 
usd_to_inr = convert_usd_to_inr()
usdt_to_usd = convert_usdt_to_usd()
eth_to_inr = convert_eth_to_inr()
#eth_to_usdt = convert_eth_to_usdt()
cryptofile = os.getcwd() + "/" + sys.argv[1]
print (cryptofile)
csvfile = os.getcwd() + '/arbitrage_data.csv'
mail_message = []
headers = ['CURRENT_DATE_AND_TIME','CUR','BUY SOURCE','BUYPRICE','SELL SOURCE','SELLPRICE','% DIFFERENCE', 'FUND']
emailid_sendto = 'sample@example.com'
emailid_sendfrom = 'sample2@example.com'
password_sendfrom = 'sendfrompassword'



def current_date_and_time():
	return datetime.datetime.now().strftime('%d/%b/%Y: %I:%M %p')


def roundoff_float(var):
	return round(var,3)


def send_mail(sender_addr,receiver_addr,message,subject,password):
	msg = MIMEMultipart()
	msg['From'] = sender_addr
	msg['To'] = receiver_addr
	msg['Subject'] = subject
	msg.attach(MIMEText(message))

	s = smtplib.SMTP('smtp.gmail.com',587)
	s.starttls()
	s.login(sender_addr,password)
	s.sendmail(sender_addr,receiver_addr,msg.as_string())
	s.quit()


def create_and_save(row):
	global mail_message
	global headers
	#store for mail_message creation
	mail_message.append(row)

	
	if os.path.exists(csvfile):
		append_write = 'a' # append if already exists
	else:
		append_write = 'w' # make a new file if not
	
	
	#add headers to csv file
	if len(mail_message) <= 1:
		with open(csvfile, append_write) as f:
			f.write(",".join(str(e) for e in headers) + "\n")
		with open(csvfile, 'a') as f:
			f.write(",".join(str(e) for e in row) + "\n")
	else:
		with open(csvfile, append_write) as f:
			f.write(",".join(str(e) for e in row) + "\n")




def get_price(url,path,conversion):
	try:
		r = requests.get(url)
		data = json.loads(r.text)
	except:
		print (url,path,conversion)
		exit(1)
	
	indexes = path.split(",")

	for index in indexes:
		if type(data) is list:
			data = data[int(index)] 
		else:
			data = data[index]

	data = float(data)
	if conversion is not None:
		data = eval("data " + conversion)

	return data
	

def calculate_arbitrage(currency, buy_list, sell_list):
	date_and_time = current_date_and_time()
	arbitrage_alert = False
	reverse_arbitrage_alert = False
	global mail_message
	mail_message = []
	
	for buy_item in buy_list:
		for sell_item in sell_list:
			arbitrage = (sell_item["price"] - buy_item["price"])/buy_item["price"]*100
			reverse_arbitrage = (buy_item["price"] - sell_item["price"])/sell_item["price"]*100
			buy_name = buy_item["name"]
			buy_price = buy_item["price"]
			sell_name = sell_item["name"]
			sell_price = sell_item["price"]

			create_and_save ([date_and_time,currency,buy_name,roundoff_float(buy_price),sell_name,roundoff_float(sell_price),roundoff_float(arbitrage),roundoff_float(reverse_arbitrage)])
			
			if arbitrage >= 25:
				arbitrage_alert = True
			if reverse_arbitrage >= 0:
				reverse_arbitrage_alert = True
	
	mail_message.sort(key=lambda x: float(x[6]), reverse=True)
	create_and_save([])
	mail_message = tabulate(mail_message,headers = headers)
	if arbitrage_alert:
		send_mail(emailid_sendfrom,emailid_sendto,mail_message,currency + " Arbitrage >= 25%",password_sendfrom)

	if reverse_arbitrage_alert:
		send_mail(emailid_sendfrom,emailid_sendto,mail_message,currency + " Fund: Buy at koinex",password_sendfrom)


def main():
	wb = load_workbook(filename = cryptofile)
	sheets = wb.get_sheet_names()

	for currency in sheets:
		buy_list = []
		sell_list = []
		ws = wb[currency]
		
		for row in ws.iter_rows(min_row=2):
			(buy_or_sell,name,url,path,conversion) = (row[0].value,row[1].value,row[2].value,row[3].value,row[4].value)
			price = get_price(url,path,conversion)
			
			if buy_or_sell.strip() == 'buy':
				buy_list.append({"name":name,"price":price})
			else:
				sell_list.append({"name":name,"price":price})

		
		calculate_arbitrage(currency, buy_list, sell_list)

		print (mail_message)


if __name__ == "__main__":
	main()
